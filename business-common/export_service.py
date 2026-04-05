#!/usr/bin/env python3
"""
数据导出服务 - V17.0新增
支持导出格式：
  - CSV (逗号分隔值)
  - Excel (.xlsx)
  - JSON (数组格式)

使用方法：
  from export_service import ExportService
  exporter = ExportService()
  result = exporter.export_orders(where, params, format='csv')
"""
import os
import sys
import csv
import io
import json
import logging
from datetime import datetime
from typing import List, Dict, Any, Optional

logger = logging.getLogger(__name__)

# 尝试导入Excel支持
try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    EXCEL_SUPPORTED = True
except ImportError:
    EXCEL_SUPPORTED = False
    logger.warning("openpyxl未安装，Excel导出功能不可用")


class ExportService:
    """数据导出服务"""
    
    # 导出文件存储目录
    EXPORT_DIR = os.path.join(os.path.dirname(__file__), '..', 'exports')
    
    def __init__(self):
        os.makedirs(self.EXPORT_DIR, exist_ok=True)
    
    def export_to_csv(self, data: List[Dict], filename: str, columns: List[Dict] = None) -> str:
        """
        导出为CSV格式
        
        参数:
            data: 数据列表
            filename: 文件名（不含扩展名）
            columns: 列配置 [{key: 'field_key', title: '显示名称'}, ...]
        
        返回:
            CSV文件路径
        """
        if not data:
            raise ValueError("没有数据可导出")
        
        # 确定列配置
        if columns is None:
            columns = [{'key': k, 'title': k} for k in data[0].keys()]
        
        # 生成文件路径
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"{filename}_{timestamp}.csv"
        filepath = os.path.join(self.EXPORT_DIR, filename)
        
        with open(filepath, 'w', newline='', encoding='utf-8-sig') as f:
            writer = csv.writer(f)
            
            # 写入表头
            writer.writerow([col['title'] for col in columns])
            
            # 写入数据
            for row in data:
                writer.writerow([row.get(col['key'], '') for col in columns])
        
        logger.info(f"CSV导出成功: {filepath}")
        return filepath
    
    def export_to_excel(self, data: List[Dict], filename: str, columns: List[Dict] = None, 
                        sheet_name: str = 'Sheet1') -> str:
        """
        导出为Excel格式
        
        参数:
            data: 数据列表
            filename: 文件名（不含扩展名）
            columns: 列配置 [{key: 'field_key', title: '显示名称'}, ...]
            sheet_name: 工作表名称
        
        返回:
            Excel文件路径
        """
        if not EXCEL_SUPPORTED:
            raise ImportError("openpyxl未安装，请执行: pip install openpyxl")
        
        if not data:
            raise ValueError("没有数据可导出")
        
        # 确定列配置
        if columns is None:
            columns = [{'key': k, 'title': k} for k in data[0].keys()]
        
        # 创建工作簿
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name[:31]  # Excel工作表名称最长31字符
        
        # 样式定义
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        # 写入表头
        for col_idx, col in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col['title'])
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # 写入数据
        for row_idx, row in enumerate(data, 2):
            for col_idx, col in enumerate(columns, 1):
                value = row.get(col['key'], '')
                # 处理特殊类型
                if isinstance(value, (list, dict)):
                    value = json.dumps(value, ensure_ascii=False)
                elif hasattr(value, 'isoformat'):
                    value = value.isoformat()
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # 自动调整列宽
        for col_idx, col in enumerate(columns, 1):
            max_length = len(str(col['title']))
            for row_idx in range(2, len(data) + 2):
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(max_length + 2, 50)
        
        # 生成文件路径
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"{filename}_{timestamp}.xlsx"
        filepath = os.path.join(self.EXPORT_DIR, filename)
        
        wb.save(filepath)
        logger.info(f"Excel导出成功: {filepath}")
        return filepath
    
    def export_to_json(self, data: List[Dict], filename: str) -> str:
        """
        导出为JSON格式
        
        参数:
            data: 数据列表
            filename: 文件名（不含扩展名）
        
        返回:
            JSON文件路径
        """
        if not data:
            raise ValueError("没有数据可导出")
        
        # 生成文件路径
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"{filename}_{timestamp}.json"
        filepath = os.path.join(self.EXPORT_DIR, filename)
        
        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump({
                'export_time': datetime.now().isoformat(),
                'total': len(data),
                'data': data
            }, f, ensure_ascii=False, indent=2)
        
        logger.info(f"JSON导出成功: {filepath}")
        return filepath
    
    def export_orders(self, where: str, params: List, format: str = 'csv') -> str:
        """导出订单数据"""
        from . import db
        
        sql = f"""
            SELECT o.order_no, o.order_type, o.order_status, o.pay_status,
                   o.total_amount, o.actual_amount, o.discount_amount,
                   o.user_name, o.user_phone, o.shipping_address,
                   o.tracking_no, o.logistics_company,
                   o.created_at, o.paid_at, o.shipped_at, o.completed_at
            FROM business_orders o
            WHERE {where}
            ORDER BY o.created_at DESC
        """
        
        data = db.get_all(sql, params)
        
        columns = [
            {'key': 'order_no', 'title': '订单编号'},
            {'key': 'order_type', 'title': '订单类型'},
            {'key': 'order_status', 'title': '订单状态'},
            {'key': 'pay_status', 'title': '支付状态'},
            {'key': 'total_amount', 'title': '总价'},
            {'key': 'actual_amount', 'title': '实付金额'},
            {'key': 'discount_amount', 'title': '优惠金额'},
            {'key': 'user_name', 'title': '用户姓名'},
            {'key': 'user_phone', 'title': '手机号'},
            {'key': 'shipping_address', 'title': '收货地址'},
            {'key': 'tracking_no', 'title': '物流单号'},
            {'key': 'logistics_company', 'title': '物流公司'},
            {'key': 'created_at', 'title': '下单时间'},
            {'key': 'paid_at', 'title': '支付时间'},
            {'key': 'shipped_at', 'title': '发货时间'},
            {'key': 'completed_at', 'title': '完成时间'}
        ]
        
        if format == 'excel':
            return self.export_to_excel(data, 'orders', columns, '订单数据')
        elif format == 'json':
            return self.export_to_json(data, 'orders')
        else:
            return self.export_to_csv(data, 'orders', columns)
    
    def export_members(self, where: str, params: List, format: str = 'csv') -> str:
        """导出会员数据"""
        from . import db
        
        sql = f"""
            SELECT m.user_id, m.user_name, m.phone, m.member_level,
                   m.points, m.total_points, m.balance, m.total_consume,
                   m.total_checkin_days, m.max_checkin_days,
                   m.last_checkin_date, m.created_at, m.updated_at
            FROM business_members m
            WHERE {where}
            ORDER BY m.created_at DESC
        """
        
        data = db.get_all(sql, params)
        
        columns = [
            {'key': 'user_id', 'title': '用户ID'},
            {'key': 'user_name', 'title': '用户姓名'},
            {'key': 'phone', 'title': '手机号'},
            {'key': 'member_level', 'title': '会员等级'},
            {'key': 'points', 'title': '当前积分'},
            {'key': 'total_points', 'title': '累计积分'},
            {'key': 'balance', 'title': '余额'},
            {'key': 'total_consume', 'title': '累计消费'},
            {'key': 'total_checkin_days', 'title': '累计签到'},
            {'key': 'max_checkin_days', 'title': '最大连续签到'},
            {'key': 'last_checkin_date', 'title': '最后签到日期'},
            {'key': 'created_at', 'title': '注册时间'},
            {'key': 'updated_at', 'title': '更新时间'}
        ]
        
        if format == 'excel':
            return self.export_to_excel(data, 'members', columns, '会员数据')
        elif format == 'json':
            return self.export_to_json(data, 'members')
        else:
            return self.export_to_csv(data, 'members', columns)
    
    def export_applications(self, where: str, params: List, format: str = 'csv') -> str:
        """导出会申请数据"""
        from . import db
        
        sql = f"""
            SELECT a.app_no, a.app_type, a.title, a.content, a.status,
                   a.priority, a.user_name, a.user_phone,
                   a.assignee_name, a.result, a.created_at, a.updated_at, a.completed_at
            FROM business_applications a
            WHERE {where}
            ORDER BY a.created_at DESC
        """
        
        data = db.get_all(sql, params)
        
        columns = [
            {'key': 'app_no', 'title': '申请编号'},
            {'key': 'app_type', 'title': '申请类型'},
            {'key': 'title', 'title': '标题'},
            {'key': 'content', 'title': '内容'},
            {'key': 'status', 'title': '状态'},
            {'key': 'priority', 'title': '优先级'},
            {'key': 'user_name', 'title': '申请人'},
            {'key': 'user_phone', 'title': '手机号'},
            {'key': 'assignee_name', 'title': '处理人'},
            {'key': 'result', 'title': '处理结果'},
            {'key': 'created_at', 'title': '创建时间'},
            {'key': 'updated_at', 'title': '更新时间'},
            {'key': 'completed_at', 'title': '完成时间'}
        ]
        
        if format == 'excel':
            return self.export_to_excel(data, 'applications', columns, '申请数据')
        elif format == 'json':
            return self.export_to_json(data, 'applications')
        else:
            return self.export_to_csv(data, 'applications', columns)
    
    def export_products(self, where: str, params: List, format: str = 'csv') -> str:
        """导出商品数据"""
        from . import db
        
        sql = f"""
            SELECT p.id, p.product_name, p.category, p.price, p.original_price,
                   p.stock, p.sales_count, p.view_count, p.favorite_count,
                   p.status, p.created_at, p.updated_at
            FROM business_products p
            WHERE {where}
            ORDER BY p.created_at DESC
        """
        
        data = db.get_all(sql, params)
        
        columns = [
            {'key': 'id', 'title': '商品ID'},
            {'key': 'product_name', 'title': '商品名称'},
            {'key': 'category', 'title': '分类'},
            {'key': 'price', 'title': '售价'},
            {'key': 'original_price', 'title': '原价'},
            {'key': 'stock', 'title': '库存'},
            {'key': 'sales_count', 'title': '销量'},
            {'key': 'view_count', 'title': '浏览量'},
            {'key': 'favorite_count', 'title': '收藏数'},
            {'key': 'status', 'title': '状态'},
            {'key': 'created_at', 'title': '创建时间'},
            {'key': 'updated_at', 'title': '更新时间'}
        ]
        
        if format == 'excel':
            return self.export_to_excel(data, 'products', columns, '商品数据')
        elif format == 'json':
            return self.export_to_json(data, 'products')
        else:
            return self.export_to_csv(data, 'products', columns)
    
    def get_export_url(self, filepath: str, base_url: str = '') -> str:
        """获取导出文件的下载URL"""
        filename = os.path.basename(filepath)
        return f"{base_url}/exports/{filename}" if base_url else f"/exports/{filename}"


# 便捷函数
def quick_export(data: List[Dict], export_type: str, format: str = 'csv') -> str:
    """
    快速导出数据
    
    参数:
        data: 数据列表
        export_type: 导出类型 (orders/members/applications/products)
        format: 导出格式 (csv/excel/json)
    
    返回:
        导出文件路径
    """
    exporter = ExportService()
    
    if export_type == 'orders':
        # 需要从数据库查询
        pass
    elif export_type == 'members':
        pass
    elif export_type == 'applications':
        pass
    elif export_type == 'products':
        pass
    
    if format == 'excel':
        return exporter.export_to_excel(data, export_type)
    elif format == 'json':
        return exporter.export_to_json(data, export_type)
    else:
        return exporter.export_to_csv(data, export_type)


if __name__ == '__main__':
    # 测试导出功能
    test_data = [
        {'id': 1, 'name': '测试用户', 'value': 100},
        {'id': 2, 'name': '测试用户2', 'value': 200}
    ]
    
    exporter = ExportService()
    
    print("测试CSV导出...")
    csv_file = exporter.export_to_csv(test_data, 'test')
    print(f"CSV文件: {csv_file}")
    
    if EXCEL_SUPPORTED:
        print("测试Excel导出...")
        excel_file = exporter.export_to_excel(test_data, 'test')
        print(f"Excel文件: {excel_file}")
    
    print("测试JSON导出...")
    json_file = exporter.export_to_json(test_data, 'test')
    print(f"JSON文件: {json_file}")
